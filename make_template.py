#!/usr/bin/env python3
"""Génère le template gratuit RelanceSimple (suivi de relances + mise en forme conditionnelle)."""
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

OUT = "/root/Projets/MyHeliosProduit/h3-landing/assets/relancesimple-template.xlsx"

# ---- Styles ----
DARK = "0B1120"
HEAD_FILL = PatternFill("solid", fgColor=DARK)
HEAD_FONT = Font(bold=True, color="FFFFFF", size=12)
TITLE_FONT = Font(bold=True, size=16, color="E2E8F0")
HINT_FONT = Font(italic=True, size=10, color="94A3B8")
RED_FILL = PatternFill("solid", fgColor="FDE8E8")
ORANGE_FILL = PatternFill("solid", fgColor="FEF3E2")
BLUE_FILL = PatternFill("solid", fgColor="E3F0FD")
THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

# ================= Sheet 1 : Suivi =================
ws = wb.active
ws.title = "Suivi"
ws.sheet_view.showGridLines = False

ws.merge_cells("A1:D1")
ws["A1"] = "RelanceSimple — Suivi des relances"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A2:D2")
ws["A2"] = ("Éditez à partir de la ligne 5. Colonne C = date de prochaine relance (JJ/MM/AAAA). "
            "La couleur s'applique automatiquement : rouge = en retard, orange = aujourd'hui, bleu = cette semaine.")
ws["A2"].font = HINT_FONT

headers = ["Prospect", "Dernière action", "Prochaine relance", "Notes"]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = HEAD_FONT
    c.fill = HEAD_FILL
    c.alignment = Alignment(vertical="center")
    c.border = BORDER

# Ligne d'exemple réaliste (date passée -> montre la règle "en retard")
example = ["Exemple — Société X", "Devis envoyé le 28/07 — relance douce prévue", datetime.date(2026, 8, 4), "Ligne d'exemple : supprimez-la"]
for col, v in enumerate(example, start=1):
    c = ws.cell(row=5, column=col, value=v)
    c.border = BORDER
    c.font = Font(italic=True, color="64748B")
ws.cell(row=5, column=3).number_format = "DD/MM/YYYY"

# Règles conditionnelles sur la colonne C (lignes 5 à 104)
rng = "C5:C104"
ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["TODAY()"], fill=RED_FILL, font=Font(bold=True, color="B91C1C")))
ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=["TODAY()"], fill=ORANGE_FILL, font=Font(bold=True, color="B45309")))
ws.conditional_formatting.add(rng, FormulaRule(formula=["AND($C5>=TODAY(),$C5<=TODAY()+7)"], fill=BLUE_FILL, font=Font(bold=True, color="1D4ED8")))

widths = {"A": 30, "B": 42, "C": 20, "D": 32}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"

# ================= Sheet 2 : Mode d'emploi =================
ws2 = wb.create_sheet("Mode d'emploi")
ws2.sheet_view.showGridLines = False
ws2.merge_cells("A1:B1")
ws2["A1"] = "Comment utiliser ce template"
ws2["A1"].font = TITLE_FONT

lines = [
    ("1.", "Après chaque échange avec un prospect, mettez à jour sa ligne : dernière action + date de prochaine relance."),
    ("2.", "Les couleurs vous disent où regarder : rouge = en retard (relancez maintenant), orange = aujourd'hui, bleu = cette semaine."),
    ("3.", "Rituel hebdomadaire (10 min) : traitez les lignes rouges puis oranges, et planifiez la semaine à venir."),
    ("4.", "Supprimez la ligne d'exemple (ligne 5) et ajoutez vos prospects."),
]
row = 3
for num, txt in lines:
    ws2.cell(row=row, column=1, value=num).font = Font(bold=True, color="0EA5E9")
    ws2.cell(row=row, column=2, value=txt).alignment = Alignment(wrap_text=True, vertical="top")
    row += 1

row += 1
ws2.cell(row=row, column=1, value="Légende des couleurs").font = Font(bold=True, size=12)
for i, (fill, txt) in enumerate([(RED_FILL, "En retard — la date de relance est dépassée"),
                                 (ORANGE_FILL, "Aujourd'hui — à relancer aujourd'hui"),
                                 (BLUE_FILL, "Cette semaine — à planifier")]):
    c1 = ws2.cell(row=row + 1 + i, column=1)
    c1.fill = fill
    c1.border = BORDER
    c2 = ws2.cell(row=row + 1 + i, column=2, value=txt)
ws2.column_dimensions["A"].width = 10
ws2.column_dimensions["B"].width = 90

wb.save(OUT)
print("OK ->", OUT)
